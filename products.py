import openpyxl


#print the number of products per company
def number_of_products(file_name):
    workbook = openpyxl.load_workbook(file_name)
    inventory_sheet = workbook["Sheet1"]  # a workbook has a lot of sheets
    company_products = {}
    for product_row in range(2, inventory_sheet.max_row + 1):
        company_name = inventory_sheet.cell(product_row, 4).value
        if company_name in company_products:
            company_products[company_name] += 1
        else:
            company_products[company_name] = 1
            print("adding a new supplier \n")
    print(company_products)

#print the total value of all products per company
def total_value_per_company(file_name):
    workbook = openpyxl.load_workbook(file_name)
    inventory_sheet = workbook["Sheet1"]
    total_per_company = {}
    for product_row in range(2, inventory_sheet.max_row + 1):
        company_name = inventory_sheet.cell(product_row, 4).value
        inventory = inventory_sheet.cell(product_row, 2).value
        price = inventory_sheet.cell(product_row, 3).value
        current_value = inventory * price
        if company_name in total_per_company:
            total_per_company[company_name] += current_value
        else:
            total_per_company[company_name] = 0
            print("adding a new supplier \n")
    print(total_per_company)

#print the products with inventory less or equal to 10
def limited_products(file_name):
    workbook = openpyxl.load_workbook(file_name)
    inventory_sheet = workbook["Sheet1"]
    products = {}
    for product_row in range(2, inventory_sheet.max_row + 1):
        product_id = int(inventory_sheet.cell(product_row, 1).value)
        product_inventory = int(inventory_sheet.cell(product_row, 2).value)

        if product_inventory <= 10:
            products[product_id] = product_inventory

    print(products)

#add a new column in the file that contains for each product its total value
def add_total_price_column(file_name):
    workbook = openpyxl.load_workbook(file_name)
    inventory_sheet = workbook["Sheet1"]
    inventory_sheet.cell(1, 5).value = "total_price"
    for product_row in range(2, inventory_sheet.max_row + 1):
        inventory = inventory_sheet.cell(product_row, 2).value
        price = inventory_sheet.cell(product_row, 3).value
        product_price = price * inventory
        inventory_sheet.cell(product_row, 5).value = product_price
    workbook.save(file_name)
